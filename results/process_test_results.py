import os
import re
import pandas as pd

def create_excel_for_each_seasonal_trend():
    # Base directory containing the results
    base_dir = "/home/noam.koren/multiTS/NFT/results"

    # Create an output directory for the merged Excel files
    output_dir = os.path.join(base_dir, "merged_results")
    os.makedirs(output_dir, exist_ok=True)

    # Filter directories with the specific format
    dirs = []
    for seasonal_amplitude in [0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 4, 5, 6, 7, 8]:
        for trend_amplitude in [0, 0.2, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 30, 40, 50, 60]:
            d = f'seasonal_{seasonal_amplitude}_trend_{trend_amplitude}'
            if os.path.isdir(os.path.join(base_dir, d)):
                dirs.append(d)
    print(dirs)

    # Process each directory
    for dir_name in dirs:
        dir_path = os.path.join(base_dir, dir_name)
        models = ["DLinear", "nft", "PatchTST", "TimesNet"]
        merged_data = None

        # Process each model's Excel file in the directory
        for model in models:
            model_path = os.path.join(dir_path, model)
            print(model_path)
            excel_file = os.path.join(model_path, f"{model}_{dir_name}_results.xlsx")
            if os.path.exists(excel_file) and os.path.getsize(excel_file) > 0:
                # Read the Excel file
                df = pd.read_excel(excel_file)
                
                # Skip rows with Lookback 35 or 150 for nft
                if model == "nft":
                    df = df[~df["Lookback"].isin([35, 150])]
                
                df = df.drop_duplicates(subset=["Lookback", "Horizon"], keep="first")

                # Select relevant columns
                subset_df = df[["Data", "Lookback", "Horizon", "test_mse"]].rename(columns={"test_mse": model})
                
                # Merge data
                if merged_data is None:
                    merged_data = subset_df
                else:
                    merged_data = pd.merge(merged_data, subset_df, on=["Data", "Lookback", "Horizon"], how="outer")

        # Save the merged data to a new Excel file
        if merged_data is not None:
            output_file = os.path.join(output_dir, f"{dir_name}.xlsx")
            merged_data.to_excel(output_file, index=False)

    print(f"Merged files saved in: {output_dir}")

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def consolidate_results_with_highlighting():
    # Base directory containing the results
    base_dir = "/home/noam.koren/multiTS/NFT/results"
    output_dir = os.path.join(base_dir, "merged_results")
    consolidated_file = os.path.join(output_dir, "structured_results_highlighted.xlsx")
    
    # Initialize the ranges for seasonality and trend
    seasonal_amplitudes = [0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 4, 5, 6, 7, 8]
    trend_amplitudes = [0, 0.2, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 30, 40, 50, 60]
    lookbacks = [25, 50, 100]
    horizons = [1, 10, 20, 30, 40, 50]

    # Initialize a DataFrame for the structure
    final_data = []

    for trend_amplitude in trend_amplitudes:
        for lookback in lookbacks:
            for horizon in horizons:
                row = {
                    "Trend": trend_amplitude,
                    "Lookback": lookback,
                    "Horizon": horizon
                }
                
                # Add seasonality columns for this row
                for seasonal_amplitude in seasonal_amplitudes:
                    # Prepare the column names for this seasonality
                    seasonal_columns = [
                        f"{seasonal_amplitude}_nft",
                        f"{seasonal_amplitude}_dlinear",
                        f"{seasonal_amplitude}_patchtst",
                        f"{seasonal_amplitude}_timesnet"
                    ]
                    
                    # Collect model data for this seasonality
                    model_data = []
                    for model in ["nft", "DLinear", "PatchTST", "TimesNet"]:
                        file_name = f"seasonal_{seasonal_amplitude}_trend_{trend_amplitude}.xlsx"
                        file_path = os.path.join(output_dir, file_name)
                        
                        if os.path.exists(file_path):
                            df = pd.read_excel(file_path)
                            if model in df.columns:
                                # Filter for the current Lookback and Horizon
                                filtered_df = df[(df["Lookback"] == lookback) & (df["Horizon"] == horizon)]
                                if not filtered_df.empty:
                                    model_data.append(filtered_df[model].values[0])
                                else:
                                    model_data.append(None)
                            else:
                                model_data.append(None)  # Column missing
                        else:
                            model_data.append(None)  # File missing
                    
                    # Add the model data to the row
                    for col_name, value in zip(seasonal_columns, model_data):
                        row[col_name] = value
                
                # Append the row to the final data
                final_data.append(row)

    # Convert to a DataFrame
    final_df = pd.DataFrame(final_data)

    # Save to Excel
    final_df.to_excel(consolidated_file, index=False)

    # Highlight lowest values
    highlight_lowest_values(consolidated_file, seasonal_amplitudes)
    print(f"Structured results with highlights saved in: {consolidated_file}")


def highlight_lowest_values(file_path, seasonal_amplitudes):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Get column names from the header row
    headers = [cell.value for cell in ws[1]]

    # Highlight each seasonality group
    for seasonal_amplitude in seasonal_amplitudes:
        # Find the columns for this seasonality
        seasonal_columns = [
            headers.index(f"{seasonal_amplitude}_nft") + 1,
            headers.index(f"{seasonal_amplitude}_dlinear") + 1,
            headers.index(f"{seasonal_amplitude}_patchtst") + 1,
            headers.index(f"{seasonal_amplitude}_timesnet") + 1
        ]

        # Iterate over rows to find the lowest value
        for row in range(2, ws.max_row + 1):  # Start from row 2 (data rows)
            values = [
                ws.cell(row=row, column=col).value
                for col in seasonal_columns
                if ws.cell(row=row, column=col).value is not None
            ]
            if values:
                min_value = min(values)
                for col in seasonal_columns:
                    cell = ws.cell(row=row, column=col)
                    if cell.value == min_value:
                        # Apply yellow fill
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


    # Save the workbook
    wb.save(file_path)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_cells_with_colors(file_path, seasonal_amplitudes):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Get column names from the header row
    headers = [cell.value for cell in ws[1]]

    # Define colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    light_light_blue_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")  # Light Light Blue
    light_blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Light Blue
    blue_fill = PatternFill(start_color="6699FF", end_color="6699FF", fill_type="solid")  # Blue

    # Highlight each seasonality group
    for seasonal_amplitude in seasonal_amplitudes:
        # Find the columns for this seasonality
        seasonal_columns = [
            headers.index(f"{seasonal_amplitude}_nft") + 1,
            headers.index(f"{seasonal_amplitude}_dlinear") + 1,
            headers.index(f"{seasonal_amplitude}_patchtst") + 1,
            headers.index(f"{seasonal_amplitude}_timesnet") + 1
        ]

        # Iterate over rows to apply highlighting
        for row in range(2, ws.max_row + 1):  # Start from row 2 (data rows)
            lookback = ws.cell(row=row, column=headers.index("Lookback") + 1).value

            # Identify the minimum value for yellow highlight
            values = [
                ws.cell(row=row, column=col).value
                for col in seasonal_columns
                if ws.cell(row=row, column=col).value is not None
            ]
            min_value = min(values) if values else None

            for col in seasonal_columns:
                cell = ws.cell(row=row, column=col)
                if cell.value == min_value:
                    # Highlight the lowest value in yellow
                    cell.fill = yellow_fill
                else:
                    # Apply blue shades based on Lookback
                    if lookback == 25:
                        cell.fill = light_light_blue_fill
                    elif lookback == 50:
                        cell.fill = light_blue_fill
                    elif lookback == 100:
                        cell.fill = blue_fill

    # Save the workbook
    wb.save(file_path)


create_excel_for_each_seasonal_trend()
# Call the function to consolidate and structure the results
consolidate_results_with_highlighting()
# Example usage after generating the consolidated file
consolidated_file = "/home/noam.koren/multiTS/NFT/results/merged_results/structured_results_highlighted.xlsx"
seasonal_amplitudes = [
0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 4, 5, 6, 7, 8]
highlight_cells_with_colors(consolidated_file, seasonal_amplitudes)

